import pandas as pd
from sqlalchemy import create_engine
from openpyxl import load_workbook
from openpyxl.styles import (Font, PatternFill, Alignment,
                              Border, Side)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
import datetime
from config import DB_CONFIG

engine = create_engine(
    f"postgresql+psycopg2://{DB_CONFIG['user']}:{DB_CONFIG['password']}@"
    f"{DB_CONFIG['host']}:{DB_CONFIG['port']}/{DB_CONFIG['database']}"
)

print("📊 Pulling data from PostgreSQL...")

# ── Pull all datasets ────────────────────────────────────
q_top10 = """
    SELECT keyword,
           ROUND(AVG(interest),1) AS avg_interest,
           MAX(interest)          AS peak_interest
    FROM search_trends
    WHERE interest > 0
    GROUP BY keyword
    ORDER BY avg_interest DESC
    LIMIT 10;
"""

q_category = """
    SELECT
        CASE
            WHEN keyword IN ('smartphones Kenya','laptops Kenya','TVs Kenya','earphones Kenya','generators Kenya') THEN 'Electronics'
            WHEN keyword IN ('clothes online Kenya','sneakers Kenya','handbags Kenya','segunda Kenya','mitumba Kenya') THEN 'Fashion'
            WHEN keyword IN ('supermarket Kenya','food delivery Kenya','cooking gas Kenya','unga Kenya','sukuma wiki Kenya') THEN 'Food & Grocery'
            WHEN keyword IN ('salon equipment Kenya','wigs Kenya','weaves Kenya','skincare Kenya','perfume Kenya') THEN 'Beauty & Salon'
            WHEN keyword IN ('houses for sale Nairobi','land Kenya','bedsitter Nairobi','Airbnb Kenya','rental houses Kenya') THEN 'Real Estate'
            WHEN keyword IN ('fertilizer Kenya','seeds Kenya','poultry Kenya','greenhouse Kenya','dairy farming Kenya') THEN 'Agriculture'
            WHEN keyword IN ('loans Kenya','Mpesa Kenya','Fuliza Kenya','forex Kenya','crypto Kenya') THEN 'Finance'
            WHEN keyword IN ('pharmacy Kenya','hospital Kenya','insurance Kenya','chemist Kenya','dawa Kenya') THEN 'Health'
        END AS category,
        ROUND(AVG(interest),1) AS avg_interest,
        MAX(interest)          AS peak_interest,
        COUNT(*)               AS data_points
    FROM search_trends
    WHERE interest > 0
    GROUP BY category
    ORDER BY avg_interest DESC;
"""

q_monthly = """
    SELECT
        TO_CHAR(date, 'Mon YYYY')         AS month,
        DATE_TRUNC('month', date)         AS month_date,
        ROUND(AVG(interest), 1)           AS avg_interest,
        COUNT(DISTINCT keyword)           AS keywords_tracked
    FROM search_trends
    WHERE interest > 0
    GROUP BY TO_CHAR(date, 'Mon YYYY'), DATE_TRUNC('month', date)
    ORDER BY month_date ASC;
"""

q_rising = """
    SELECT
        keyword,
        ROUND(AVG(CASE WHEN date >= CURRENT_DATE - INTERVAL '3 months'
              THEN interest END), 1) AS recent_avg,
        ROUND(AVG(CASE WHEN date < CURRENT_DATE - INTERVAL '3 months'
              AND date >= CURRENT_DATE - INTERVAL '6 months'
              THEN interest END), 1) AS previous_avg
    FROM search_trends
    WHERE interest > 0
    GROUP BY keyword
    HAVING
        AVG(CASE WHEN date >= CURRENT_DATE - INTERVAL '3 months'
            THEN interest END) IS NOT NULL
        AND
        AVG(CASE WHEN date < CURRENT_DATE - INTERVAL '3 months'
            AND date >= CURRENT_DATE - INTERVAL '6 months'
            THEN interest END) IS NOT NULL
    ORDER BY recent_avg DESC
    LIMIT 10;
"""

q_raw = """
    SELECT date, keyword, interest, region
    FROM search_trends
    ORDER BY date ASC, keyword ASC;
"""

df_top10    = pd.read_sql(q_top10,    engine)
df_category = pd.read_sql(q_category, engine)
df_monthly  = pd.read_sql(q_monthly,  engine)
df_rising   = pd.read_sql(q_rising,   engine)
df_raw      = pd.read_sql(q_raw,      engine)

df_top10["keyword"]    = df_top10["keyword"].str.replace(" Kenya","").str.replace(" Nairobi","")
df_rising["keyword"]   = df_rising["keyword"].str.replace(" Kenya","").str.replace(" Nairobi","")
df_rising["change"]    = df_rising["recent_avg"] - df_rising["previous_avg"]
df_monthly             = df_monthly.drop(columns=["month_date"])

print("✅ Data pulled successfully!")

# ── Write to Excel ───────────────────────────────────────
file_path = "outputs/Kenya_Demand_Analysis.xlsx"

with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
    df_top10.to_excel(writer,    sheet_name="Top 10 Products",  index=False)
    df_category.to_excel(writer, sheet_name="By Category",      index=False)
    df_monthly.to_excel(writer,  sheet_name="Monthly Trend",    index=False)
    df_rising.to_excel(writer,   sheet_name="Rising vs Falling",index=False)
    df_raw.to_excel(writer,      sheet_name="Raw Data",         index=False)

print("✅ Sheets written!")

# ── Style the workbook ───────────────────────────────────
wb = load_workbook(file_path)

# Colours
DARK_BLUE  = "1F3864"
GOLD       = "C9A84C"
LIGHT_BLUE = "D9E2F3"
WHITE      = "FFFFFF"
GREEN      = "375623"
RED        = "C00000"

def style_sheet(ws, title):
    """Apply professional styling to a worksheet."""

    # ── Header row styling ───────────────────────────────
    for cell in ws[1]:
        cell.font      = Font(bold=True, color=WHITE, size=11)
        cell.fill      = PatternFill("solid", fgColor=DARK_BLUE)
        cell.alignment = Alignment(horizontal="center",
                                   vertical="center", wrap_text=True)
        cell.border    = Border(
            bottom=Side(style="medium", color=GOLD)
        )

    # ── Data rows ────────────────────────────────────────
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill_color = LIGHT_BLUE if i % 2 == 0 else WHITE
        for cell in row:
            cell.fill      = PatternFill("solid", fgColor=fill_color)
            cell.alignment = Alignment(horizontal="center",
                                       vertical="center")
            cell.border    = Border(
                bottom=Side(style="thin", color="CCCCCC")
            )

    # ── Column widths ────────────────────────────────────
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 6

    # ── Row height ───────────────────────────────────────
    ws.row_dimensions[1].height = 30

    # ── Title above table ────────────────────────────────
    ws.insert_rows(1)
    ws.merge_cells(f"A1:{get_column_letter(ws.max_column)}1")
    title_cell            = ws["A1"]
    title_cell.value      = title
    title_cell.font       = Font(bold=True, color=WHITE, size=13)
    title_cell.fill       = PatternFill("solid", fgColor=DARK_BLUE)
    title_cell.alignment  = Alignment(horizontal="center",
                                      vertical="center")
    ws.row_dimensions[1].height = 35


# Apply styling to all sheets
style_sheet(wb["Top 10 Products"],
            "🇰🇪 Kenya — Top 10 Most Searched Products (Jun 2025 – Jun 2026)")
style_sheet(wb["By Category"],
            "🇰🇪 Kenya — Consumer Demand by Category")
style_sheet(wb["Monthly Trend"],
            "🇰🇪 Kenya — Monthly Demand Trend")
style_sheet(wb["Rising vs Falling"],
            "🇰🇪 Kenya — Rising vs Falling Products")
style_sheet(wb["Raw Data"],
            "🇰🇪 Kenya — Raw Google Trends Data")

# ── Highlight change column on Rising sheet ──────────────
ws_rising = wb["Rising vs Falling"]
change_col = 4  # column D

for row in ws_rising.iter_rows(min_row=3, min_col=change_col,
                                max_col=change_col):
    for cell in row:
        if cell.value is not None:
            if float(cell.value) > 0:
                cell.font = Font(bold=True, color=GREEN)
            elif float(cell.value) < 0:
                cell.font = Font(bold=True, color=RED)

wb.save(file_path)
engine.dispose()

print(f"\n🎉 Excel report saved!")
print(f"📁 Location: outputs/Kenya_Demand_Analysis.xlsx")
print(f"📋 Sheets: Top 10 Products | By Category | Monthly Trend | Rising vs Falling | Raw Data")